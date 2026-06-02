from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models.match import MatchPhase
from app.schemas.match import MatchCreate
from app.schemas.team import TeamCreate
from app.schemas.tournament import TournamentCreate
from app.services.match_service import create_match
from app.services.team_service import (
    create_team,
    get_team_by_name_or_code,
    list_teams,
)
from app.services.tournament_service import (
    create_tournament,
    list_tournaments,
)


REQUIRED_COLUMNS = [
    "torneo",
    "anio",
    "fase",
    "grupo",
    "local",
    "codigo_local",
    "visitante",
    "codigo_visitante",
    "fecha_partido",
    "cierre_prediccion",
]


PHASE_ALIASES: dict[str, MatchPhase] = {
    "FASE DE GRUPOS": MatchPhase.GROUP_STAGE,
    "GRUPOS": MatchPhase.GROUP_STAGE,
    "GROUP_STAGE": MatchPhase.GROUP_STAGE,

    "16AVOS": MatchPhase.ROUND_OF_32,
    "DIECISEISAVOS": MatchPhase.ROUND_OF_32,
    "ROUND_OF_32": MatchPhase.ROUND_OF_32,

    "OCTAVOS": MatchPhase.ROUND_OF_16,
    "ROUND_OF_16": MatchPhase.ROUND_OF_16,

    "CUARTOS": MatchPhase.QUARTER_FINAL,
    "QUARTER_FINAL": MatchPhase.QUARTER_FINAL,

    "SEMIFINAL": MatchPhase.SEMI_FINAL,
    "SEMIFINALES": MatchPhase.SEMI_FINAL,
    "SEMI_FINAL": MatchPhase.SEMI_FINAL,

    "TERCER PUESTO": MatchPhase.THIRD_PLACE,
    "TERCER_PUESTO": MatchPhase.THIRD_PLACE,
    "THIRD_PLACE": MatchPhase.THIRD_PLACE,

    "FINAL": MatchPhase.FINAL,
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def normalize_code(value: Any) -> str:
    return normalize_text(value).upper()


def normalize_phase(value: Any) -> MatchPhase | None:
    raw = normalize_text(value).upper()

    if not raw:
        return None

    raw = raw.replace("-", " ").replace("_", " ")
    raw = " ".join(raw.split())

    if raw in PHASE_ALIASES:
        return PHASE_ALIASES[raw]

    raw_underscore = raw.replace(" ", "_")

    return PHASE_ALIASES.get(raw_underscore)


def parse_year(value: Any) -> int | None:
    if value is None or value == "":
        return None

    try:
        return int(value)
    except Exception:
        return None


def parse_excel_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)

        return value.astimezone(timezone.utc)

    raw = str(value).strip()

    formats = [
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue

    try:
        parsed = datetime.fromisoformat(raw)
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except ValueError:
        return None


def create_fixture_template_excel() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "fixture"

    ws.append(REQUIRED_COLUMNS)

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    sample_rows = [
        [
            "Mundial 2026",
            2026,
            "Fase de grupos",
            "A",
            "Argentina",
            "ARG",
            "Francia",
            "FRA",
            "2026-06-20 18:00",
            "2026-06-20 17:00",
        ],
        [
            "Mundial 2026",
            2026,
            "Fase de grupos",
            "A",
            "Brasil",
            "BRA",
            "Alemania",
            "GER",
            "2026-06-21 15:00",
            "2026-06-21 14:00",
        ],
    ]

    for row in sample_rows:
        ws.append(row)

    widths = {
        "A": 22,
        "B": 10,
        "C": 20,
        "D": 12,
        "E": 22,
        "F": 16,
        "G": 22,
        "H": 18,
        "I": 22,
        "J": 22,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    instructions = wb.create_sheet("instrucciones")
    instructions.append(["Campo", "Descripción"])
    instructions.append(["torneo", "Nombre del torneo. Ejemplo: Mundial 2026"])
    instructions.append(["anio", "Año del torneo. Ejemplo: 2026"])
    instructions.append(["fase", "Fase de grupos, 16avos, Octavos, Cuartos, Semifinal, Tercer puesto, Final"])
    instructions.append(["grupo", "Grupo del Mundial. Ejemplo: A, B, C. Puede quedar vacío en eliminatorias"])
    instructions.append(["local", "Nombre del equipo local"])
    instructions.append(["codigo_local", "Código corto del equipo local. Ejemplo: ARG"])
    instructions.append(["visitante", "Nombre del equipo visitante"])
    instructions.append(["codigo_visitante", "Código corto del equipo visitante. Ejemplo: FRA"])
    instructions.append(["fecha_partido", "Formato recomendado: YYYY-MM-DD HH:MM"])
    instructions.append(["cierre_prediccion", "Debe ser anterior o igual a fecha_partido"])

    for cell in instructions[1]:
        cell.fill = header_fill
        cell.font = header_font

    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 90

    output = BytesIO()
    wb.save(output)

    return output.getvalue()


async def read_uploaded_workbook(file: UploadFile):
    filename = file.filename or ""

    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser Excel .xlsx o .xlsm",
        )

    content = await file.read()

    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo está vacío",
        )

    try:
        return load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se pudo leer el archivo Excel",
        )


def get_fixture_sheet(workbook):
    if "fixture" in workbook.sheetnames:
        return workbook["fixture"]

    return workbook.active


def read_excel_rows(workbook) -> list[dict[str, Any]]:
    ws = get_fixture_sheet(workbook)

    headers = []
    for cell in ws[1]:
        headers.append(normalize_text(cell.value).lower())

    missing = [column for column in REQUIRED_COLUMNS if column not in headers]

    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Faltan columnas obligatorias: {', '.join(missing)}",
        )

    header_index = {name: index for index, name in enumerate(headers)}

    rows = []

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None or str(value).strip() == "" for value in row):
            continue

        item = {"row_number": row_number}

        for column in REQUIRED_COLUMNS:
            item[column] = row[header_index[column]] if header_index[column] < len(row) else None

        rows.append(item)

    return rows


def validate_import_row(raw: dict[str, Any]) -> dict[str, Any]:
    errors: list[str] = []

    torneo = normalize_text(raw.get("torneo"))
    anio = parse_year(raw.get("anio"))
    fase_raw = normalize_text(raw.get("fase"))
    phase = normalize_phase(fase_raw)
    grupo = normalize_text(raw.get("grupo")).upper() or None

    local = normalize_text(raw.get("local"))
    codigo_local = normalize_code(raw.get("codigo_local"))
    visitante = normalize_text(raw.get("visitante"))
    codigo_visitante = normalize_code(raw.get("codigo_visitante"))

    fecha_partido = parse_excel_datetime(raw.get("fecha_partido"))
    cierre_prediccion = parse_excel_datetime(raw.get("cierre_prediccion"))

    if not torneo:
        errors.append("Falta torneo")

    if anio is None:
        errors.append("Año inválido")

    if phase is None:
        errors.append("Fase inválida")

    if not local:
        errors.append("Falta equipo local")

    if not codigo_local:
        errors.append("Falta código local")

    if not visitante:
        errors.append("Falta equipo visitante")

    if not codigo_visitante:
        errors.append("Falta código visitante")

    if codigo_local and codigo_visitante and codigo_local == codigo_visitante:
        errors.append("El equipo local y visitante no pueden ser el mismo")

    if fecha_partido is None:
        errors.append("Fecha de partido inválida")

    if cierre_prediccion is None:
        errors.append("Fecha de cierre inválida")

    if fecha_partido and cierre_prediccion and cierre_prediccion > fecha_partido:
        errors.append("El cierre de predicción no puede ser posterior al partido")

    return {
        "row_number": raw["row_number"],
        "torneo": torneo,
        "anio": anio,
        "fase": fase_raw,
        "phase": phase,
        "grupo": grupo,
        "local": local,
        "codigo_local": codigo_local,
        "visitante": visitante,
        "codigo_visitante": codigo_visitante,
        "fecha_partido": fecha_partido,
        "cierre_prediccion": cierre_prediccion,
        "valid": len(errors) == 0,
        "errors": errors,
    }


def build_preview_response(validated_rows: list[dict[str, Any]]) -> dict[str, Any]:
    rows = []

    for row in validated_rows:
        rows.append(
            {
                "row_number": row["row_number"],
                "torneo": row["torneo"] or None,
                "anio": row["anio"],
                "fase": row["fase"] or None,
                "grupo": row["grupo"],
                "local": row["local"] or None,
                "codigo_local": row["codigo_local"] or None,
                "visitante": row["visitante"] or None,
                "codigo_visitante": row["codigo_visitante"] or None,
                "fecha_partido": row["fecha_partido"].isoformat() if row["fecha_partido"] else None,
                "cierre_prediccion": row["cierre_prediccion"].isoformat() if row["cierre_prediccion"] else None,
                "valid": row["valid"],
                "errors": row["errors"],
            }
        )

    total_rows = len(rows)
    valid_rows = len([row for row in rows if row["valid"]])
    invalid_rows = total_rows - valid_rows

    return {
        "ok": invalid_rows == 0,
        "total_rows": total_rows,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "rows": rows,
    }


async def preview_fixture_import(file: UploadFile) -> dict[str, Any]:
    workbook = await read_uploaded_workbook(file)
    raw_rows = read_excel_rows(workbook)
    validated_rows = [validate_import_row(row) for row in raw_rows]

    return build_preview_response(validated_rows)


def find_or_create_tournament(db: Session, name: str, year: int):
    tournaments = list_tournaments(db)

    for tournament in tournaments:
        if tournament.name.strip().lower() == name.strip().lower() and tournament.year == year:
            return tournament

    return create_tournament(
        db,
        TournamentCreate(
            name=name,
            year=year,
            description=f"Torneo importado desde Excel: {name} {year}",
            is_active=True,
        ),
    )


def find_or_create_team(
    db: Session,
    name: str,
    code: str,
    created_counter: dict[str, int],
):
    existing = get_team_by_name_or_code(db, name, code)

    if existing:
        return existing

    team = create_team(
        db,
        TeamCreate(
            name=name,
            code=code,
            flag_url=None,
        ),
    )

    created_counter["count"] += 1

    return team


async def commit_fixture_import(db: Session, file: UploadFile) -> dict[str, Any]:
    workbook = await read_uploaded_workbook(file)
    raw_rows = read_excel_rows(workbook)
    validated_rows = [validate_import_row(row) for row in raw_rows]

    errors: list[str] = []

    invalid_rows = [row for row in validated_rows if not row["valid"]]

    if invalid_rows:
        for row in invalid_rows:
            errors.append(f"Fila {row['row_number']}: {', '.join(row['errors'])}")

        return {
            "ok": False,
            "total_rows": len(validated_rows),
            "imported_matches": 0,
            "created_teams": 0,
            "skipped_rows": len(invalid_rows),
            "errors": errors,
        }

    created_teams_counter = {"count": 0}
    imported_matches = 0

    for row in validated_rows:
        try:
            tournament = find_or_create_tournament(
                db,
                name=row["torneo"],
                year=row["anio"],
            )

            home_team = find_or_create_team(
                db,
                name=row["local"],
                code=row["codigo_local"],
                created_counter=created_teams_counter,
            )

            away_team = find_or_create_team(
                db,
                name=row["visitante"],
                code=row["codigo_visitante"],
                created_counter=created_teams_counter,
            )

            create_match(
                db,
                MatchCreate(
                    tournament_id=tournament.id,
                    phase=row["phase"],
                    world_group=row["grupo"],
                    home_team_id=home_team.id,
                    away_team_id=away_team.id,
                    match_datetime=row["fecha_partido"],
                    prediction_deadline=row["cierre_prediccion"],
                ),
            )

            imported_matches += 1

        except Exception as exc:
            errors.append(f"Fila {row['row_number']}: {str(exc)}")

    return {
        "ok": len(errors) == 0,
        "total_rows": len(validated_rows),
        "imported_matches": imported_matches,
        "created_teams": created_teams_counter["count"],
        "skipped_rows": len(errors),
        "errors": errors,
    }